'''
    @Author: Ly Nguyen & Nathan Lucyk
    @Date: 2.4.2014

    This script is used to upload the abstracts in json format to s3, and users and abstract links to dynamodb. Currently this script does not work with non-ASCII encodings.
    However, certain encodings like utf-8 and unicode may still work.


    This script is unreliable and should be refactored and improved. The script does not work for non-english words (no accents allowed!), 
    and makes some assumptions about the data which may not be true such as the presentation number being an integer or that it exists at all.
    To make this fault more tolerable, the index of whatever failed is printed to a .txt file and logged in a log file. 

    Timings were done while uploading large amounts of users to dynamo and abstracts to s3 and found that a large majority of the work is the upload
    process.

    Future extensions:
        Get multiple encodings to work to allow more data to be added. 
        Get rid of globals. 
        Having a whitelist of attributes we do accept into dynamo / s3. 
        Way to automagically convert fieldnames into acceptable (uniform from whitelist) formats (E.G., CONTACTEMAIL -> email)

'''

import sys
import logging
import csv
import json 
import argparse
import time
from boto.s3.key import Key
from boto.dynamodb2.table import Table
from boto import s3
from boto import dynamodb2
from openpyxl import load_workbook

SEQUENCER_TABLE_NAME = 'Sequencer'
USER_TABLE_NAME = 'User_Test'
PAPER_TABLE_NAME = 'Paper_Test'
ABSTRACT_BUCKET_NAME = 'sagebionetworks-interactome-abstracts'
PRESENTATION_NUM_EXCEL_FIELD = 'PresentationNumber'
LASTNAME_EXCEL_FIELD = 'LastName'
FIRSTNAME_EXCEL_FIELD = 'FirstName'
# These are defined by the credentials csv downloaded from amazon
SECRET_KEY_EXCEL_FIELD = 'Secret Access Key'
ACCESS_KEY_EXCEL_FIELD = 'Access Key Id'
# Tells how to format the logging and what file to store it in
logging.basicConfig(filename='dataLoader.log', level=logging.INFO, format='%(asctime)s -- %(levelname)s: %(message)s')

''' 
    Parses the csv file for ACCESS_KEY_EXCEL_FIELD and SECRET_KEY_EXCEL_FIELD.
    It will then use those credentials to connect to s3, dynamo, and all tables.
    This does not check if the keys have enough authority to add to these resources (this may be a good addition in the future).
'''
def connectToAWS(authCSVFileName):
    REGION = 'us-west-2'
    # These resources are used throughout the application.
    global s3Conn
    global dynamoConn
    global sequencerTable
    global usersTable
    global papersTable
    global abstractToPaperDict
    # Practically every line could throw an exception so we use one try
    try:
    # Opens the credentials and then uses the first line. Using DictReader may be overkill
        csvfile = open(authCSVFileName, 'rU') 
        reader = csv.DictReader(csvfile) 
        authLine = reader.next()
        accessKey = authLine[ACCESS_KEY_EXCEL_FIELD]
        secretKey = authLine[SECRET_KEY_EXCEL_FIELD]
        # Get S3 and dynamo connection
        s3Conn = s3.connect_to_region(REGION, aws_access_key_id=accessKey, aws_secret_access_key=secretKey)
        dynamoConn = dynamodb2.connect_to_region(REGION, aws_access_key_id=accessKey, aws_secret_access_key=secretKey)
        # Use connections to load tables
        sequencerTable = Table(SEQUENCER_TABLE_NAME, connection=dynamoConn)
        usersTable = Table(USER_TABLE_NAME, connection=dynamoConn)
        papersTable = Table(PAPER_TABLE_NAME, connection=dynamoConn)
    except Exception, e:
        print(e)
        sys.exit(1)

'''
    Grabs the sequence from the sequence table. If it fails, it will return None.
        Note: This should retry if failing for an access reason. If it fails for some other weird reason it will return None.
'''
def getSequence():
    sequenceNumber = None
    try:
      sequenceNumber = dynamoConn.update_item(SEQUENCER_TABLE_NAME, 
                                    {"Id":{"N":"1"}}, 
                                    attribute_updates={"Sequence":{"Action":"ADD", "Value":{"N":"1"}}}, 
                                    expected=None, 
                                    return_values="UPDATED_NEW", 
                                    return_consumed_capacity=None, 
                                    return_item_collection_metrics=None)['Attributes']['Sequence']['N']
    except Exception, e:
        print(e)
    return sequenceNumber

'''
    Parses the csv file and adds the users found within to the users table. 
        Warning: This does not check for dupe users.
'''
def addUsers(excelFileName):
    WRITES_ALLOWED_PER_LOOP = 32
    # Load the excel stuff. See http://pythonhosted.org/openpyxl/api.html#module-openpyxl-reader-iter-worksheet-optimized-reader
    workbook = load_workbook(filename = excelFileName, use_iterators = True)
    first_sheet = workbook.get_sheet_names()[0]
    worksheet = workbook.get_sheet_by_name(first_sheet)

    lostIndexFile = open("indexesNotAddedUSERS.txt", 'w')

    rowIndex = 0

    for row in worksheet.iter_rows():
        rowIndex += 1
        # skip row with header info. (Don't use iter to make sure we only read one row at a time into memory)
        if(rowIndex == 1):
            continue
        # put to sleep for a second so that we don't go over the throughput allowed
        elif(rowIndex % WRITES_ALLOWED_PER_LOOP == 0):
            time.sleep(1)
        try:
            #control number + presentation number
            hashKey = str(row[2].internal_value) + str(row[3].internal_value)     
            firstname = row[4].internal_value
            lastname = row[5].internal_value
            institution = row[6].internal_value
            email = row[7].internal_value
            

            # query to see if user was already added
            userId = ''
            userQueryResults = usersTable.query_2(Email__eq=email, index='Email-index')
            listUserQueryResults = list(userQueryResults)
            # We check all results due to moving institutions, changing names, and such.
            for user in listUserQueryResults:
                if(user['Institution'] == institution and user['LastName'] == lastname and user['FirstName'] == firstname):
                    userId = user['Id']
                    break
            try:
                if(userId == ''):

                    sequenceNumber = getSequence()
                    if (sequenceNumber == None):
                        lostIndexFile.write(str(rowIndex) + "\n")
                        logging.error("Unabled to sequence item. Index not added: " + str(rowIndex))
                        continue
                    userId = "User" + str(sequenceNumber)


                    newItem = { 'Id': userId,
                        'LastName': lastname,
                        'FirstName': firstname,
                        'Email': email,
                        'Institution': institution
                    }
                    usersTable.put_item(data=newItem)
                elif(hashKey in abstractToPaperDict):
                    # update user to have paperId
                    dynamoPaperId = abstractToPaperDict[hashKey]
                    userItem = usersTable.get_item(hash_key=userId)
                    if('Papers' in userItem):
                        userItem['Papers'].add(dynamoPaperId)
                    else:
                        userItem['Papers'] = set([dynamoPaperId])
                    userItem.save()

                    # Update paper to have author
                    paperItem = paperTable.get_item(hash_key=dynamoPaperId)
                    if('Authors' in paperItem):
                        paperItem['Authors'].add(userId)
                    else:
                        paperItem['Authors'] = set([userId])
                    paperItem.save()

                else:
                    lostIndexFile.write(str(rowIndex) + "\n")
                    logging.error("Unabled to find hashkey in dict. Index not added: " + str(rowIndex))
            except:
                lostIndexFile.write(str(rowIndex) + "\n")
                logging.error("Unabled to add to usersTable. Index not added: " + str(rowIndex))
        except Exception, e:
            lostIndexFile.write(str(rowIndex) + "\n")
            logging.error("Unknown abstract error. Index not added: " + str(rowIndex) + "\nErr: " + str(e))

    lostIndexFile.close()

'''
    Parses the csv file for abstracts. It will query to find if the author is already a user. If so, puts the user id with the abstract
    The abstract will be dumped into S3 in json format. The URL will be stored in the papers table.
        Warning: does not check for dupes
'''
def addAbstracts(excelFileName):
    WRITES_ALLOWED_PER_LOOP = 32
    abstractToPaperDict = {}

    # Load the excel stuff. See http://pythonhosted.org/openpyxl/api.html#module-openpyxl-reader-iter-worksheet-optimized-reader
    workbook = load_workbook(filename = excelFileName, use_iterators = True)
    first_sheet = workbook.get_sheet_names()[0]
    worksheet = workbook.get_sheet_by_name(first_sheet)

    lostIndexFile = open("indexesNotAddedABSTRACTS.txt", 'w')
    abstractsBucket = s3Conn.get_bucket(ABSTRACT_BUCKET_NAME)
    abstractKey = Key(abstractsBucket)
    rowIndex = 0

    for row in worksheet.iter_rows():
        rowIndex += 1
        # skip row with header info. (Don't use iter to make sure we only read one row at a time into memory)
        if(rowIndex == 1):
            continue
        # put to sleep for a second so that we don't go over the throughput allowed
        elif(rowIndex % WRITES_ALLOWED_PER_LOOP == 0):
            time.sleep(1)
        try:
            #control number + presentation number
            print row[0].internal_value
            print row[1].internal_value
            hashKey = str(row[0].internal_value) + str(row[1].internal_value)
            print hashKey
            return
            s3Format = None
            abstractTitle = ""
            try:
                abstractTitle = row[2].internal_value.encode('utf-8')
                s3Format = {"AbstractTitle": abstractTitle,
                    "Abstract": row[3].internal_value.encode('utf-8'),
                    "Subclass": row[4].internal_value.encode('utf-8'),
                    "Category-Subcat-Subclass": row[5].internal_value.encode('utf-8'),
                    "Keyword1": row[6].internal_value.encode('utf-8'),
                    "Keyword2": row[7].internal_value.encode('utf-8'),
                    "Keyword3": row[8].internal_value.encode('utf-8'),
                    "Keyword4": row[9].internal_value.encode('utf-8'),
                    "SessionTitle": row[10].internal_value.encode('utf-8'),
                    "PresenterFirstname": row[11].internal_value.encode('utf-8'),
                    "PresenterLastname": row[12].internal_value.encode('utf-8'),
                    "PresenterInstitution": row[13].internal_value.encode('utf-8'),
                    "PresenterEmail": row[14].internal_value.encode('utf-8')}
            except:
                 lostIndexFile.write(str(rowIndex) + "\n")
                 logging.error("S3Formatting failed. Index not added: " + str(rowIndex))
                 continue

            sequenceNumber = getSequence()
            if (sequenceNumber == None):
                lostIndexFile.write(str(rowIndex) + "\n")
                logging.error("Unabled to sequence item. Index not added: " + str(rowIndex))
                continue

            abstractKey.key = str(rowIndex) + 'Abstract' + str(rowIndex) + '.json'
            abstractKey.set_metadata("Content-Type", 'application/json')
            abstractKey.set_contents_from_string(s3Format)
            abstractKey.make_public()
            abstractUrlLink = abstractKey.generate_url(0, query_auth=False, force_http=True)

            # Now link the S3 url to an entry in the dynamodb table
            dynamoPaperId = 'PaperTEST' + str(rowIndex) # + sequenceNumber
            newItem = {'Id': dynamoPaperId, 'Link': abstractUrlLink, 'Title': abstractTitle}
            try:
                papersTable.put_item(data=newItem)
            except:
                lostIndexFile.write(str(rowIndex) + "\n")
                logging.error("Unabled to add to papers table. Index not added: " + str(rowIndex))
            else:
                abstractToPaperDict[hashKey] = dynamoPaperId

        except Exception, e:
            lostIndexFile.write(str(rowIndex) + "\n")
            logging.error("Unknown abstract error. Index not added: " + str(rowIndex) + "\nErr: " + str(e))
    print "abstractsToPapers! ------------------- \n"
    print abstractToPaperDict
    lostIndexFile.close()


def main(argv=sys.argv):
    parser = argparse.ArgumentParser(description='Process optional download flags')

    AUTH_HELP_TEXT = "path to credential csv file from Amazon. This will not be stored. CSV fieldnames must be \"" \
                        + SECRET_KEY_EXCEL_FIELD + "\" for secret key and \"" + ACCESS_KEY_EXCEL_FIELD +"\" for access key"
    parser.add_argument('--auth', dest='usersAuth', help=AUTH_HELP_TEXT, required=True)
    parser.add_argument('--users', dest='usersPath', help='path to users csv file', default=argparse.SUPPRESS)
    parser.add_argument('--abstracts', dest='abstractsPath', help='path to abstracts csv file', default=argparse.SUPPRESS)
    options = parser.parse_args(argv[1:])

    connectToAWS(options.usersAuth)

    if hasattr(options, 'usersPath'):
        addUsers(options.usersPath)
    
    if hasattr(options, 'abstractsPath'):
        addAbstracts(options.abstractsPath)
        

if __name__ == "__main__":
    sys.exit(main())
